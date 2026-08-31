#!/usr/bin/env python3
from __future__ import annotations

import argparse,csv,hashlib,io,json,math,re,zipfile
from datetime import date,datetime,time
from pathlib import Path
from typing import Any,Iterable

EXPECTED_HEADER=["日付","時間","始値","高値","安値","終値","出来高"]
TRANSFORM_VERSION="JNU_MACD_4223_POSTPUB_G1_V1"

def sha256_file(path:Path)->str:
    h=hashlib.sha256()
    with path.open("rb") as fh:
        for b in iter(lambda:fh.read(1024*1024),b""): h.update(b)
    return h.hexdigest()

def normalize_date(v:Any)->date:
    if isinstance(v,datetime): return v.date()
    if isinstance(v,date): return v
    s=str(v).strip().replace("/","-")
    try:return datetime.fromisoformat(s).date()
    except:return datetime.strptime(s.split()[0],"%Y-%m-%d").date()

def normalize_time(v:Any)->time:
    if isinstance(v,datetime): return v.time().replace(microsecond=0)
    if isinstance(v,time): return v.replace(microsecond=0)
    if isinstance(v,(int,float)):
        sec=round(float(v)*86400)%86400
        return time(sec//3600,(sec%3600)//60,sec%60)
    return time.fromisoformat(str(v).strip()).replace(microsecond=0)

def minute_of_day(t:time)->int:return t.hour*60+t.minute
def hhmm(s:str)->int:
    h,m=map(int,s.replace("+1","").split(":")); return h*60+m

def schedule_for(d:date,cal:dict[str,Any])->dict[str,Any]:
    ds=d.isoformat()
    for r in cal["ose_nikkei_index_futures"]:
        if ds<r["valid_from"]: continue
        if r["valid_to"] is not None and ds>r["valid_to"]: continue
        return r
    raise ValueError(ds)

def workbook_payload(path:Path):
    if path.suffix.lower() in {".xls",".xlsx"}:
        return path.name,path.read_bytes(),path.suffix.lower()
    with zipfile.ZipFile(path) as zf:
        books=[i for i in zf.infolist() if not i.is_dir() and Path(i.filename).suffix.lower() in {".xls",".xlsx"}]
        if len(books)!=1: raise ValueError(path.name)
        i=books[0]; return i.filename,zf.read(i),Path(i.filename).suffix.lower()

def sheet_names(raw:bytes,suf:str):
    if suf==".xlsx":
        import openpyxl
        wb=openpyxl.load_workbook(io.BytesIO(raw),read_only=True,data_only=True)
        try:return list(wb.sheetnames)
        finally:wb.close()
    import xlrd
    b=xlrd.open_workbook(file_contents=raw,on_demand=True)
    try:return list(b.sheet_names())
    finally:b.release_resources()

def iter_xlsx(raw:bytes,sh:str)->Iterable[tuple[Any,...]]:
    import openpyxl
    wb=openpyxl.load_workbook(io.BytesIO(raw),read_only=True,data_only=True)
    try: yield from wb[sh].iter_rows(values_only=True)
    finally: wb.close()

def iter_xls(raw:bytes,sh:str)->Iterable[tuple[Any,...]]:
    import xlrd
    b=xlrd.open_workbook(file_contents=raw,on_demand=True)
    try:
        s=b.sheet_by_name(sh)
        for r in range(s.nrows):
            vals=[]
            for c in range(s.ncols):
                cell=s.cell(r,c)
                vals.append(xlrd.xldate.xldate_as_datetime(cell.value,b.datemode) if cell.ctype==xlrd.XL_CELL_DATE else cell.value)
            yield tuple(vals)
    finally:b.release_resources()

def find_header(rows,source,sh):
    it=iter(rows)
    for _ in range(12):
        row=next(it,None)
        if row is None: break
        vals=[str(x).strip() if x is not None else "" for x in list(row)[:7]]
        if vals==EXPECTED_HEADER:return it
    raise ValueError(f"{source}/{sh}:header")

def source_files(folder:Path,product:str)->dict[int,Path]:
    files=(list(folder.glob("N225minif_*.zip"))+list(folder.glob("225mini20*d.xls"))) if product=="MINI" else list(folder.glob("N225microf_*.zip"))
    out={}
    for p in files:
        m=re.search(r"(20\d{2})",p.name)
        if m:out[int(m.group(1))]=p
    return out

def parse_days(path:Path,start:date,end:date):
    member,raw,suf=workbook_payload(path)
    one=[n for n in sheet_names(raw,suf) if str(n).strip().startswith("1min")]
    if not one: raise ValueError(f"{path.name}:no1min")
    days={}; dup=invalid=0
    for sh in one:
        rows=iter_xlsx(raw,sh) if suf==".xlsx" else iter_xls(raw,sh)
        it=find_header(rows,path.name,sh)
        for row in it:
            if not row or all(v in (None,"") for v in row): continue
            try:d=normalize_date(row[0])
            except:continue
            if d<start or d>end:continue
            try:t=normalize_time(row[1]); c=float(row[5])
            except: invalid+=1; continue
            if not(c>0 and math.isfinite(c)):invalid+=1;continue
            m=minute_of_day(t)
            rec=days.setdefault(d,{})
            if m in rec: dup+=1; continue
            rec[m]=c
    return days,{"workbook_member":member,"sheets":one,"duplicate_rows":dup,"invalid_rows":invalid,"distinct_dates":len(days)}

def ema(values:list[float],span:int)->list[float]:
    a=2/(span+1)
    out=[]
    cur=None
    for x in values:
        cur=x if cur is None else a*x+(1-a)*cur
        out.append(cur)
    return out

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--product",choices=["MINI","MICRO"],required=True)
    ap.add_argument("--input-dir",type=Path,required=True)
    ap.add_argument("--calendar",type=Path,required=True)
    ap.add_argument("--prereg",type=Path,required=True)
    ap.add_argument("--output-dir",type=Path,required=True)
    ap.add_argument("--parser-commit",required=True)
    args=ap.parse_args()
    args.output_dir.mkdir(parents=True,exist_ok=True)
    cal=json.loads(args.calendar.read_text(encoding="utf-8"))
    pre=json.loads(args.prereg.read_text(encoding="utf-8"))
    st="stage_a" if args.product=="MINI" else "stage_b"
    cfg=pre["data"][st]
    if args.product=="MINI":
        source_from=date.fromisoformat(cfg["warmup_from"]); eval_from=date.fromisoformat(cfg["evaluation_from"]); eval_to=date.fromisoformat(cfg["evaluation_to"])
    else:
        source_from=date.fromisoformat(cfg["source_from"]); eval_from=None; eval_to=date.max
    cov_min=float(pre["data_quality"]["day_session_minute_label_coverage_minimum"])
    files=source_files(args.input_dir,args.product)
    source_hash={}; all_days={}; sources=[]; critical=[]
    for year,p in sorted(files.items()):
        if year<source_from.year-1 or year>eval_to.year+1:continue
        dd,meta=parse_days(p,source_from,eval_to)
        digest=sha256_file(p); source_hash[p.name]=digest
        sources.append({"source_id":p.name,"nominal_year":year,"sha256":digest,"meta_1m":meta})
        if meta["invalid_rows"]>0: critical.append({"year":year,"issue":"INVALID_ROWS","count":meta["invalid_rows"]})
        for d,bars in dd.items():
            if d not in all_days or year==d.year:all_days[d]=(p,bars)
    daily=[]; excluded=[]
    for d,(p,bars) in sorted(all_days.items()):
        sched=schedule_for(d,cal)
        mins=[]
        for a,b in sched["day_session_segments"]:mins.extend(range(hhmm(a),hhmm(b)))
        cov=sum(1 for m in mins if m in bars)/len(mins) if mins else 0
        if cov<cov_min or mins[-1] not in bars:
            excluded.append({"trading_date":d.isoformat(),"reason":"COVERAGE_OR_FINAL_CLOSE_MISSING","coverage":cov});continue
        daily.append({"date":d,"close":bars[mins[-1]],"coverage":cov,"source":p.name})
    closes=[x["close"] for x in daily]
    e4=ema(closes,4); e22=ema(closes,22)
    macd=[a-b for a,b in zip(e4,e22)]
    sig=ema(macd,3)
    hist=[m-s for m,s in zip(macd,sig)]
    rows=[]
    burn=66 if args.product=="MICRO" else 0
    for i in range(len(daily)-1):
        d=daily[i]["date"]
        eligible=(d>=eval_from and d<=eval_to) if args.product=="MINI" else (i>=burn)
        if not eligible:continue
        h=hist[i]; pos=1 if h>0 else (-1 if h<0 else 0)
        nxt=math.log(daily[i+1]["close"]/daily[i]["close"])
        payoff=pos*nxt if pos!=0 else 0.0
        acc_den=1 if pos!=0 and nxt!=0 else 0
        correct=1 if acc_den and ((nxt>0 and pos>0) or (nxt<0 and pos<0)) else 0
        rows.append({
            "signal_date":d.isoformat(),"next_date":daily[i+1]["date"].isoformat(),
            "histogram":h,"position":pos,"next_day_return":nxt,"signal_payoff":payoff,
            "accuracy_denominator":acc_den,"correct":correct,
            "day_session_minute_coverage":daily[i]["coverage"],
            "source_file_sha256":source_hash[daily[i]["source"]],"transform_version":TRANSFORM_VERSION
        })
    stem="jnu_macd_4223_mini_stage_a_g1" if args.product=="MINI" else "jnu_macd_4223_micro_stage_b_g1"
    panel=args.output_dir/f"{stem}.csv"
    fields=["signal_date","next_date","histogram","position","next_day_return","signal_payoff","accuracy_denominator","correct","day_session_minute_coverage","source_file_sha256","transform_version"]
    with panel.open("w",encoding="utf-8",newline="") as fh:
        w=csv.DictWriter(fh,fieldnames=fields,lineterminator="\n");w.writeheader();w.writerows(rows)
    manifest={
      "version":"1.0","candidate_id":pre["candidate_id"],"stage":"A_TRUE_OSE_MINI" if args.product=="MINI" else "B_EXACT_JNU_MICRO",
      "source_license_classification":"225LABO_PERSONAL_USE_LOCAL_RAW_DERIVED_NON_RECONSTRUCTIVE_EXPORT",
      "raw_data_cloud_uploaded":False,"parser_version_commit":args.parser_commit,"calendar_session_version":cal.get("version"),
      "product_contract_coverage":{"venue":"OSE","product":"Nikkei 225 Mini Futures" if args.product=="MINI" else "Nikkei 225 Micro Futures (JNU)"},
      "date_range":[rows[0]["signal_date"] if rows else None,rows[-1]["signal_date"] if rows else None],
      "missingness_summary":{"usable_signal_days":len(rows),"excluded_trading_days":len(excluded),"coverage_gate":cov_min},
      "duplicate_summary":{"source_duplicate_minute_rows":sum(s["meta_1m"]["duplicate_rows"] for s in sources)},
      "derived_feature_definitions":{
        "histogram":"EMA4(close)-EMA22(close)-EMA3(EMA4-EMA22)",
        "position":"sign(histogram_t)",
        "next_day_return":"log(day_close[t+1]/day_close[t])",
        "signal_payoff":"position_t*next_day_return"
      },
      "source_hashes":[{"source_id":s["source_id"],"sha256":s["sha256"]} for s in sources],
      "sources":sources,"excluded_days":excluded[:200],"critical_data_quality_issues":critical,
      "derived_output_hash":sha256_file(panel),"market_outcome_interpretation_performed":False
    }
    mp=args.output_dir/f"{stem}_manifest.json";mp.write_text(json.dumps(manifest,ensure_ascii=False,indent=2)+"\n",encoding="utf-8")
    print(json.dumps({"usable_days":len(rows),"critical":len(critical),"date_range":manifest["date_range"],"panel_sha256":manifest["derived_output_hash"],"panel":str(panel),"manifest":str(mp)},ensure_ascii=False,indent=2))
if __name__=="__main__":main()
