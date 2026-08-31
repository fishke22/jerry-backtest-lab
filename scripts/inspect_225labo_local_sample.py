#!/usr/bin/env python3
"""Fail-closed local schema auditor for user-downloaded 225Labo files.

This tool is intentionally Phase-A only:
- never logs credentials/cookies/authenticated URLs;
- never uploads or prints market-data rows;
- keeps raw files local;
- emits only file hashes, structural schema metadata, coverage candidates,
  and non-reconstructive diagnostics.

It does NOT implement the HAR-RSV transform and does NOT infer a concrete
225Labo parser from an unseen file.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import re
import zipfile
from pathlib import Path
from typing import Any, Iterable

SUPPORTED_TEXT = {".csv", ".txt"}
SUPPORTED_EXCEL = {".xlsx", ".xlsm"}
SUPPORTED_ARCHIVE = {".zip"}

HEADER_HINTS = {
    "date", "datetime", "time", "open", "high", "low", "close", "volume",
    "日付", "日時", "時刻", "始値", "高値", "安値", "終値", "出来高",
}


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for block in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(block)
    return h.hexdigest()


def looks_like_header(row: list[str]) -> bool:
    vals = [str(x).strip() for x in row if str(x).strip()]
    if not vals:
        return False
    lowered = {v.lower() for v in vals}
    if lowered & HEADER_HINTS:
        return True
    # Header rows are usually mostly non-numeric labels.
    non_numeric = 0
    for v in vals:
        try:
            float(v.replace(",", ""))
        except Exception:
            non_numeric += 1
    return non_numeric / max(1, len(vals)) >= 0.7


def safe_header(row: list[Any]) -> list[str] | None:
    vals = ["" if x is None else str(x).strip() for x in row]
    return vals if looks_like_header(vals) else None


def detect_delimiter(sample: str) -> str:
    try:
        return csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
    except Exception:
        return ","


def inspect_text_bytes(raw: bytes, source_name: str) -> dict[str, Any]:
    text = None
    encoding = None
    for enc in ("utf-8-sig", "cp932", "shift_jis", "utf-8"):
        try:
            text = raw.decode(enc)
            encoding = enc
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        return {"name": source_name, "status": "FAIL_UNSUPPORTED_TEXT_ENCODING"}

    sample = text[:16384]
    delim = detect_delimiter(sample)
    reader = csv.reader(io.StringIO(text), delimiter=delim)
    rows = []
    for idx, row in enumerate(reader):
        rows.append(row)
        if idx >= 49:
            break

    result: dict[str, Any] = {
        "name": source_name,
        "status": "STRUCTURE_INSPECTED",
        "encoding": encoding,
        "delimiter": "\\t" if delim == "\t" else delim,
        "sampled_row_count": len(rows),
        "column_count_first_nonempty": next((len(r) for r in rows if r), 0),
        "header_detected": False,
        "header": None,
    }
    if rows:
        hdr = safe_header(rows[0])
        if hdr is not None:
            result["header_detected"] = True
            result["header"] = hdr
    return result


def inspect_excel_path(path: Path) -> dict[str, Any]:
    try:
        import openpyxl  # type: ignore
    except Exception as exc:
        return {
            "name": path.name,
            "status": "FAIL_OPENPYXL_NOT_AVAILABLE",
            "detail": type(exc).__name__,
        }

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        first_rows = []
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            first_rows.append(list(row))
            if idx >= 9:
                break
        hdr = safe_header(first_rows[0]) if first_rows else None
        sheets.append(
            {
                "sheet": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "header_detected": hdr is not None,
                "header": hdr,
            }
        )
    wb.close()
    return {"name": path.name, "status": "STRUCTURE_INSPECTED", "sheets": sheets}


def inspect_xlsx_bytes(raw: bytes, source_name: str) -> dict[str, Any]:
    try:
        import openpyxl  # type: ignore
    except Exception as exc:
        return {
            "name": source_name,
            "status": "FAIL_OPENPYXL_NOT_AVAILABLE",
            "detail": type(exc).__name__,
        }
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        first_rows = []
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            first_rows.append(list(row))
            if idx >= 9:
                break
        hdr = safe_header(first_rows[0]) if first_rows else None
        sheets.append(
            {
                "sheet": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "header_detected": hdr is not None,
                "header": hdr,
            }
        )
    wb.close()
    return {"name": source_name, "status": "STRUCTURE_INSPECTED", "sheets": sheets}


def inspect_zip(path: Path) -> dict[str, Any]:
    members: list[dict[str, Any]] = []
    with zipfile.ZipFile(path) as zf:
        for info in zf.infolist():
            if info.is_dir():
                continue
            suffix = Path(info.filename).suffix.lower()
            entry: dict[str, Any] = {
                "name": info.filename,
                "size": info.file_size,
                "compressed_size": info.compress_size,
                "suffix": suffix,
            }
            if suffix in SUPPORTED_TEXT:
                entry["structure"] = inspect_text_bytes(zf.read(info), info.filename)
            elif suffix in SUPPORTED_EXCEL:
                entry["structure"] = inspect_xlsx_bytes(zf.read(info), info.filename)
            else:
                entry["structure"] = {"status": "NOT_INSPECTED_UNSUPPORTED_MEMBER_TYPE"}
            members.append(entry)
    return {"archive_member_count": len(members), "members": members}


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("path", type=Path)
    ap.add_argument("--output", type=Path, default=None)
    ap.add_argument("--product", choices=["MINI", "MICRO"], default=None)
    ap.add_argument("--expected-interval", default="5m")
    args = ap.parse_args()

    path = args.path.resolve()
    if not path.is_file():
        raise SystemExit(f"225LABO_SCHEMA_AUDIT_FAIL: file not found: {path}")

    suffix = path.suffix.lower()
    report: dict[str, Any] = {
        "schema_version": "1.0",
        "audit_type": "225LABO_LOCAL_PHASE_A_SCHEMA_AUDIT",
        "raw_data_cloud_uploaded": False,
        "source_path_basename": path.name,
        "source_sha256": sha256_file(path),
        "source_size_bytes": path.stat().st_size,
        "extension": suffix,
        "declared_product": args.product,
        "expected_interval": args.expected_interval,
        "parser_status": "NOT_IMPLEMENTED_PENDING_SCHEMA_AUDIT",
        "raw_rows_emitted": False,
    }

    if suffix in SUPPORTED_ARCHIVE:
        report["structure"] = inspect_zip(path)
    elif suffix in SUPPORTED_TEXT:
        report["structure"] = inspect_text_bytes(path.read_bytes(), path.name)
    elif suffix in SUPPORTED_EXCEL:
        report["structure"] = inspect_excel_path(path)
    else:
        report["structure"] = {"status": "FAIL_UNSUPPORTED_FILE_TYPE"}

    output = args.output or path.with_name(path.stem + ".schema_manifest.json")
    output.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"225LABO_SCHEMA_AUDIT_WRITTEN={output}")
    print(f"SOURCE_SHA256={report['source_sha256']}")
    print("RAW_ROWS_EMITTED=false")


if __name__ == "__main__":
    main()
