#!/usr/bin/env python3
"""Local-only BOJ MPM event-volatility adapter for 225Labo OSE data.

Reads personally licensed 1-minute Mini/Micro workbooks and emits only
per-event non-reconstructive realized-volatility aggregates. Raw bars never
leave the local machine.

Frozen G1 windows (minute-start labels):
- baseline returns: t in [release-40m, release-10m) -> 30 one-minute returns
- event returns:    t in [release-10m, release+20m) -> 30 one-minute returns

For return r_t = log(C_t / C_{t-1}), computing both windows requires contiguous
closes from release-41 through release+19 inclusive.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import math
import re
import zipfile
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable

EXPECTED_HEADER=["日付","時間","始値","高値","安値","終値","出来高"]
EPS=1e-18
TRANSFORM_VERSION="BOJ_MPM_EVENT_VOL_G1_V1"


def sha256_file(path: Path)->str:
    h=hashlib.sha256()
    with path.open("rb") as fh:
        for b in iter(lambda:fh.read(1024*1024),b""):
            h.update(b)
    return h.hexdigest()


def normalize_date(v:Any)->date:
    if isinstance(v,datetime): return v.date()
    if isinstance(v,date): return v
    s=str(v).strip().replace("/","-")
    try: return datetime.fromisoformat(s).date()
    except ValueError: return datetime.strptime(s.split()[0],"%Y-%m-%d").date()


def normalize_time(v:Any)->time:
    if isinstance(v,datetime): return v.time().replace(microsecond=0)
    if isinstance(v,time): return v.replace(microsecond=0)
    if isinstance(v,(int,float)):
        sec=round(float(v)*86400)%86400
        return time(sec//3600,(sec%3600)//60,sec%60)
    return time.fromisoformat(str(v).strip()).replace(microsecond=0)


def minute_of_day(v:time)->int:
    return v.hour*60+v.minute


def iter_xlsx(raw:bytes,sheet:str)->Iterable[tuple[Any,...]]:
    import openpyxl
    wb=openpyxl.load_workbook(io.BytesIO(raw),read_only=True,data_only=True)
    try:
        ws=wb[sheet]
        yield from ws.iter_rows(values_only=True)
    finally:
        wb.close()


def iter_xls(raw:bytes,sheet:str)->Iterable[tuple[Any,...]]:
    import xlrd
    book=xlrd.open_workbook(file_contents=raw,on_demand=True)
    try:
        sh=book.sheet_by_name(sheet)
        for r in range(sh.nrows):
            vals=[]
            for c in range(sh.ncols):
                cell=sh.cell(r,c)
                if cell.ctype==xlrd.XL_CELL_DATE:
                    vals.append(xlrd.xldate.xldate_as_datetime(cell.value,book.datemode))
                else:
                    vals.append(cell.value)
            yield tuple(vals)
    finally:
        book.release_resources()


def workbook_payload(path:Path)->tuple[str,bytes,str]:
    suf=path.suffix.lower()
    if suf in {".xls",".xlsx"}:
        return path.name,path.read_bytes(),suf
    if suf!=".zip":
        raise ValueError(f"unsupported source type: {path}")
    with zipfile.ZipFile(path) as zf:
        books=[i for i in zf.infolist() if not i.is_dir() and Path(i.filename).suffix.lower() in {".xls",".xlsx"}]
        if len(books)!=1:
            raise ValueError(f"{path.name}: expected exactly one workbook, got {len(books)}")
        info=books[0]
        return info.filename,zf.read(info),Path(info.filename).suffix.lower()


def sheet_names(raw:bytes,suffix:str)->list[str]:
    if suffix==".xlsx":
        import openpyxl
        wb=openpyxl.load_workbook(io.BytesIO(raw),read_only=True,data_only=True)
        try: return list(wb.sheetnames)
        finally: wb.close()
    import xlrd
    b=xlrd.open_workbook(file_contents=raw,on_demand=True)
    try: return list(b.sheet_names())
    finally: b.release_resources()


def find_header(rows:Iterable[tuple[Any,...]],source:str,sheet:str)->Iterable[tuple[Any,...]]:
    it=iter(rows)
    pre=[]
    for _ in range(12):
        try: row=next(it)
        except StopIteration: break
        vals=[str(x).strip() if x is not None else "" for x in list(row)[:7]]
        if vals==EXPECTED_HEADER:
            return it
        pre.append(vals)
    raise ValueError(f"{source}/{sheet}: standard header not found; preamble={pre[:3]}")


def source_files(folder:Path,product:str)->dict[int,Path]:
    if product=="MINI":
        files=list(folder.glob("N225minif_*.zip"))+list(folder.glob("225mini20*d.xls"))
    else:
        files=list(folder.glob("N225microf_*.zip"))
    out={}
    for p in files:
        m=re.search(r"(20\d{2})",p.name)
        if m: out[int(m.group(1))]=p
    return out


def collect_event_closes(path:Path,event_dates:set[date])->tuple[dict[date,dict[int,float]],dict[str,Any]]:
    member,raw,suf=workbook_payload(path)
    names=sheet_names(raw,suf)
    one=[n for n in names if str(n).strip().startswith("1min")]
    if not one:
        raise ValueError(f"{path.name}: no 1min sheet")
    # Legacy XLS generations split annual 1-minute history across multiple
    # sheets such as "1min", "1min (2)", "1min(3)", etc. Always concatenate
    # every 1min-prefixed shard and dedupe by minute. This is a data-integrity
    # correction only; event windows/model gates are unchanged.
    data={d:{} for d in event_dates}
    duplicate=0
    invalid=0
    total_relevant=0
    for sh in one:
        rows=iter_xlsx(raw,sh) if suf==".xlsx" else iter_xls(raw,sh)
        it=find_header(rows,path.name,sh)
        for row in it:
            if not row or all(v in (None,"") for v in row): continue
            try: d=normalize_date(row[0])
            except Exception: continue
            if d not in data: continue
            total_relevant+=1
            try:
                t=normalize_time(row[1])
                c=float(row[5])
            except Exception:
                invalid+=1
                continue
            if not (c>0 and math.isfinite(c)):
                invalid+=1
                continue
            m=minute_of_day(t)
            if m in data[d]:
                duplicate+=1
                continue
            data[d][m]=c
    return data,{
        "workbook_member":member,
        "sheets":one,
        "relevant_rows":total_relevant,
        "duplicate_minute_rows":duplicate,
        "invalid_close_rows":invalid,
    }


def event_features(closes:dict[int,float],release_minute:int)->tuple[dict[str,Any]|None,dict[str,Any]|None]:
    needed=list(range(release_minute-41,release_minute+20))
    missing=[m for m in needed if m not in closes]
    if missing:
        return None,{"issue":"MISSING_REQUIRED_MINUTE_LABELS","missing_count":len(missing),"first_missing":missing[:10]}
    # Return labels t correspond to bar t and use C_t/C_{t-1}.
    returns={}
    for t in range(release_minute-40,release_minute+20):
        returns[t]=math.log(closes[t]/closes[t-1])
    base=[returns[t] for t in range(release_minute-40,release_minute-10)]
    event=[returns[t] for t in range(release_minute-10,release_minute+20)]
    if len(base)!=30 or len(event)!=30:
        return None,{"issue":"RETURN_COUNT_FAIL","baseline":len(base),"event":len(event)}
    brv=float(sum(r*r for r in base))
    erv=float(sum(r*r for r in event))
    effect=float(math.log((erv+EPS)/(brv+EPS)))
    return {
        "baseline_rv_1m":brv,
        "event_rv_1m":erv,
        "log_event_to_baseline_rv_ratio":effect,
        "baseline_return_count":30,
        "event_return_count":30,
    },None


def main()->None:
    ap=argparse.ArgumentParser()
    ap.add_argument("--product",choices=["MINI","MICRO"],required=True)
    ap.add_argument("--input-dir",type=Path,required=True)
    ap.add_argument("--events",type=Path,required=True)
    ap.add_argument("--output-dir",type=Path,required=True)
    ap.add_argument("--parser-commit",required=True)
    args=ap.parse_args()
    args.output_dir.mkdir(parents=True,exist_ok=True)

    em=json.loads(args.events.read_text(encoding="utf-8"))
    stage="A_TRUE_OSE_MINI" if args.product=="MINI" else "B_EXACT_JNU_MICRO"
    selected=[e for e in em["eligible_events"] if e["stage"]==stage]
    files=source_files(args.input_dir,args.product)
    wanted_by_year={}
    for e in selected:
        wanted_by_year.setdefault(int(e["date"][:4]),set()).add(date.fromisoformat(e["date"]))

    data_by_date={}
    sources=[]
    critical=[]
    for year,dates in sorted(wanted_by_year.items()):
        p=files.get(year)
        if p is None:
            critical.append({"year":year,"issue":"MISSING_ANNUAL_SOURCE_FILE"})
            continue
        dd,meta=collect_event_closes(p,dates)
        data_by_date.update(dd)
        sources.append({
            "source_id":p.name,
            "nominal_year":year,
            "sha256":sha256_file(p),
            "size_bytes":p.stat().st_size,
            "meta_1m":meta,
        })
        if meta["invalid_close_rows"]>0:
            critical.append({"year":year,"issue":"INVALID_CLOSE_ROWS","count":meta["invalid_close_rows"]})

    rows=[]
    unusable=[]
    for e in selected:
        d=date.fromisoformat(e["date"])
        hh,mm=map(int,e["release_time_jst"].split(":"))
        rmin=hh*60+mm
        feat,err=event_features(data_by_date.get(d,{}),rmin)
        if err:
            unusable.append({"date":e["date"],"release_time_jst":e["release_time_jst"],**err})
            continue
        rows.append({
            "event_date":e["date"],
            "release_time_jst":e["release_time_jst"],
            **feat,
            "source_event_url":e["source_url"],
            "transform_version":TRANSFORM_VERSION,
        })

    stem="jnu_boj_mpm_mini_event_volatility_g1" if args.product=="MINI" else "jnu_boj_mpm_micro_event_volatility_g1"
    panel=args.output_dir/f"{stem}.csv"
    fields=[
        "event_date","release_time_jst","baseline_rv_1m","event_rv_1m",
        "log_event_to_baseline_rv_ratio","baseline_return_count","event_return_count",
        "source_event_url","transform_version"
    ]
    with panel.open("w",encoding="utf-8",newline="") as fh:
        w=csv.DictWriter(fh,fieldnames=fields,lineterminator="\n")
        w.writeheader(); w.writerows(rows)

    manifest={
        "version":"1.0",
        "candidate_id":"BOJ_MPM_TRUE_OSE_EVENT_VOLATILITY_G1",
        "stage":stage,
        "product":"OSE Nikkei 225 Mini Futures" if args.product=="MINI" else "OSE Nikkei 225 Micro Futures (JNU)",
        "source_license_classification":"225LABO_PERSONAL_USE_LOCAL_RAW_DERIVED_NON_RECONSTRUCTIVE_EXPORT",
        "raw_data_cloud_uploaded":False,
        "parser_version_commit":args.parser_commit,
        "event_manifest_sha256":sha256_file(args.events),
        "timing_eligible_event_count":len(selected),
        "usable_event_count":len(rows),
        "unusable_event_count":len(unusable),
        "unusable_events":unusable,
        "calendar_session_version":"1.2",
        "product_contract_coverage":{
            "venue":"OSE",
            "product":"Nikkei 225 Mini Futures" if args.product=="MINI" else "Nikkei 225 Micro Futures (JNU)",
            "stage":stage
        },
        "date_range":[rows[0]["event_date"] if rows else None,rows[-1]["event_date"] if rows else None],
        "missingness_summary":{
            "timing_eligible_event_count":len(selected),
            "usable_event_count":len(rows),
            "unusable_event_count":len(unusable)
        },
        "duplicate_summary":{
            "duplicate_minute_rows":sum(int(s["meta_1m"]["duplicate_minute_rows"]) for s in sources)
        },
        "derived_feature_definitions":{
            "baseline_rv_1m":"sum squared 1m log returns in [-40m,-10m)",
            "event_rv_1m":"sum squared 1m log returns in [-10m,+20m)",
            "log_event_to_baseline_rv_ratio":"log((event_rv_1m+1e-18)/(baseline_rv_1m+1e-18))"
        },
        "measurement":{
            "frequency":"source-provided 1min",
            "baseline_window":"[-40,-10)",
            "event_window":"[-10,+20)",
            "return_count_each":30,
            "effect":"log((event RV + 1e-18)/(baseline RV + 1e-18))"
        },
        "source_hashes":[{"source_id":s["source_id"],"sha256":s["sha256"]} for s in sources],
        "sources":sources,
        "critical_data_quality_issues":critical,
        "derived_output_hash":sha256_file(panel),
        "market_outcome_interpretation_performed":False,
    }
    mp=args.output_dir/f"{stem}_manifest.json"
    mp.write_text(json.dumps(manifest,ensure_ascii=False,indent=2)+"\n",encoding="utf-8")
    print(json.dumps({
        "status":"DERIVED_PANEL_BUILT" if not critical else "DERIVED_PANEL_BUILT_WITH_CRITICAL_DQ",
        "product":args.product,
        "timing_eligible":len(selected),
        "usable":len(rows),
        "unusable":len(unusable),
        "critical":len(critical),
        "panel":str(panel),
        "manifest":str(mp),
        "panel_sha256":manifest["derived_output_hash"],
    },ensure_ascii=False,indent=2))


if __name__=="__main__":
    main()
