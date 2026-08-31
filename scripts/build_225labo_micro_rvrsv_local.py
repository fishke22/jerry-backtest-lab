#!/usr/bin/env python3
"""Local-only 225Labo Nikkei 225 Micro minute -> non-reconstructive RV/RSV adapter.

Raw 225Labo files remain local. This script reads annual ZIPs containing XLS/XLSX,
validates the frozen OSE historical session calendar, and exports only daily
realized-variance / semivariance features plus a provenance manifest.

Primary measurement: source-provided 5min sheet.
Measurement QA only: source-provided 1min sheet summary.

No strategy optimization or parameter search is performed here.
"""
from __future__ import annotations

import argparse, csv, hashlib, io, json, math, re, zipfile, statistics
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable


EXPECTED_HEADER = ["日付","時間","始値","高値","安値","終値","出来高"]

def sha256_file(path: Path) -> str:
    h=hashlib.sha256()
    with path.open("rb") as f:
        for b in iter(lambda:f.read(1024*1024),b""): h.update(b)
    return h.hexdigest()

def load_calendar(path: Path) -> dict[str,Any]:
    return json.loads(path.read_text(encoding="utf-8"))

def parse_hhmm(s: str) -> int:
    base=s.replace("+1","")
    hh,mm=map(int,base.split(":"))
    return hh*60+mm

def schedule_for(d: date, cal: dict[str,Any]) -> dict[str,Any]:
    ds=d.isoformat()
    for row in cal["ose_nikkei_index_futures"]:
        if ds < row["valid_from"]: continue
        if row["valid_to"] is not None and ds > row["valid_to"]: continue
        return row
    raise ValueError(f"no session calendar version for {ds}")

def classify_session(t: time, sched: dict[str,Any]) -> str|None:
    m=t.hour*60+t.minute
    for idx,(a,b) in enumerate(sched.get("day_session_segments",[])):
        aa,bb=parse_hhmm(a),parse_hhmm(b)
        if aa <= m <= bb: return f"DAY{idx+1}"
    for idx,(a,b) in enumerate(sched.get("night_session_segments",[])):
        aa,bb=parse_hhmm(a),parse_hhmm(b)
        if "+1" in b:
            if m >= aa or m <= bb: return f"NIGHT{idx+1}"
        elif aa <= m <= bb: return f"NIGHT{idx+1}"
    return None

def iter_xlsx_rows(raw: bytes, sheet: str) -> Iterable[tuple[Any,...]]:
    import openpyxl
    wb=openpyxl.load_workbook(io.BytesIO(raw),read_only=True,data_only=True)
    try:
        ws=wb[sheet]
        yield from ws.iter_rows(values_only=True)
    finally:
        wb.close()

def iter_xls_rows(raw: bytes, sheet: str) -> Iterable[tuple[Any,...]]:
    import xlrd
    book=xlrd.open_workbook(file_contents=raw,on_demand=True)
    try:
        sh=book.sheet_by_name(sheet)
        for r in range(sh.nrows):
            vals=[]
            for c in range(sh.ncols):
                cell=sh.cell(r,c)
                if cell.ctype==xlrd.XL_CELL_DATE:
                    vals.append(xlrd.xldate.xldate_as_datetime(cell.value,book.datemode))
                else:
                    vals.append(cell.value)
            yield tuple(vals)
    finally:
        book.release_resources()

def normalize_date(v: Any) -> date:
    if isinstance(v,datetime): return v.date()
    if isinstance(v,date): return v
    if isinstance(v,(int,float)):
        # xlrd dates should already have been converted; numeric date here is unsupported.
        raise ValueError(f"unexpected numeric date {v}")
    text=str(v).strip().replace("/", "-")
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        return datetime.strptime(text.split()[0], "%Y-%m-%d").date()

def normalize_time(v: Any) -> time:
    if isinstance(v,datetime): return v.time().replace(microsecond=0)
    if isinstance(v,time): return v.replace(microsecond=0)
    if isinstance(v,(int,float)):
        # Excel fractional day.
        seconds=round(float(v)*86400)%86400
        return time(seconds//3600,(seconds%3600)//60,seconds%60)
    text=str(v).strip()
    try:
        return time.fromisoformat(text).replace(microsecond=0)
    except ValueError:
        return datetime.fromisoformat(text).time().replace(microsecond=0)

@dataclass
class DailyAcc:
    returns: list[float]
    day_returns: list[float]
    night_returns: list[float]
    valid_bars: int=0
    invalid_bars: int=0
    duplicate_keys: int=0
    session_counts: dict[str,int]|None=None
    def __post_init__(self):
        if self.session_counts is None: self.session_counts={}

def _find_header_and_rows(rows: Iterable[tuple[Any,...]], source_name: str, sheet: str):
    """Scan a small preamble for the exact seven-column 225Labo header."""
    it=iter(rows)
    pre=[]
    for _ in range(12):
        try: r=next(it)
        except StopIteration: break
        vals=[str(x).strip() if x is not None else "" for x in list(r)[:7]]
        if vals == EXPECTED_HEADER:
            return it
        pre.append(vals)
    raise ValueError(f"{source_name}/{sheet}: standard header not found in first 12 rows; preamble={pre[:3]}")

def _workbook_payload(source_path: Path) -> tuple[str, bytes, str]:
    """Return (member/name, workbook bytes, suffix) from ZIP or direct XLS/XLSX."""
    suffix=source_path.suffix.lower()
    if suffix in {".xls",".xlsx"}:
        return source_path.name, source_path.read_bytes(), suffix
    if suffix != ".zip":
        raise ValueError(f"{source_path.name}: unsupported source type {suffix}")
    with zipfile.ZipFile(source_path) as zf:
        members=[x for x in zf.infolist() if not x.is_dir()]
        books=[x for x in members if Path(x.filename).suffix.lower() in {".xls",".xlsx"}]
        if len(books)!=1:
            raise ValueError(f"{source_path.name}: expected one workbook, found {len(books)}")
        info=books[0]
        return info.filename, zf.read(info), Path(info.filename).suffix.lower()

def _sheet_names(raw: bytes, suffix: str) -> list[str]:
    if suffix==".xlsx":
        import openpyxl
        wb=openpyxl.load_workbook(io.BytesIO(raw),read_only=True,data_only=True)
        try: return list(wb.sheetnames)
        finally: wb.close()
    import xlrd
    book=xlrd.open_workbook(file_contents=raw,on_demand=True)
    try: return list(book.sheet_names())
    finally: book.release_resources()

def _iter_sheet(raw: bytes, suffix: str, sheet: str) -> Iterable[tuple[Any,...]]:
    return iter_xlsx_rows(raw,sheet) if suffix==".xlsx" else iter_xls_rows(raw,sheet)

def parse_sheet_from_source(source_path: Path, sheet: str, cal: dict[str,Any]) -> tuple[dict[date,DailyAcc],dict[str,Any]]:
    member_name,raw,suffix=_workbook_payload(source_path)
    rows=_iter_sheet(raw,suffix,sheet)
    it=_find_header_and_rows(rows,source_path.name,sheet)

    daily: dict[date,DailyAcc]={}
    seen=set()
    prev_close: dict[tuple[date,str],float]={}
    row_count=0; invalid_ohlcv=0; out_of_session=0
    dmin=None; dmax=None
    for row in it:
        if not row or all(v in (None,"") for v in row): continue
        row_count+=1
        d=normalize_date(row[0]); t=normalize_time(row[1])
        sched=schedule_for(d,cal)
        sess=classify_session(t,sched)
        acc=daily.setdefault(d,DailyAcc([] ,[],[]))
        if sess is None:
            acc.invalid_bars+=1; out_of_session+=1; continue
        try:
            o,h,l,c,vol=[float(row[i]) for i in range(2,7)]
        except Exception:
            acc.invalid_bars+=1; invalid_ohlcv+=1; continue
        if not (h>=max(o,c,l) and l<=min(o,c,h) and vol>=0 and o>0 and h>0 and l>0 and c>0):
            acc.invalid_bars+=1; invalid_ohlcv+=1; continue
        key=(d,t)
        if key in seen:
            acc.duplicate_keys+=1
            continue
        seen.add(key)
        acc.valid_bars+=1
        acc.session_counts[sess]=acc.session_counts.get(sess,0)+1
        pkey=(d,sess)
        if pkey in prev_close:
            r=math.log(c/prev_close[pkey])
            acc.returns.append(r)
            if sess.startswith("DAY"): acc.day_returns.append(r)
            else: acc.night_returns.append(r)
        prev_close[pkey]=c
        dmin=d if dmin is None or d<dmin else dmin
        dmax=d if dmax is None or d>dmax else dmax
    meta={
        "workbook_member":member_name,
        "workbook_member_size":len(raw),
        "source_container":source_path.suffix.lower(),
        "sheet":sheet,
        "row_count":row_count,
        "distinct_trading_dates":len(daily),
        "date_min":dmin.isoformat() if dmin else None,
        "date_max":dmax.isoformat() if dmax else None,
        "invalid_ohlcv_rows":invalid_ohlcv,
        "out_of_session_rows":out_of_session,
        "duplicate_timestamp_rows":sum(a.duplicate_keys for a in daily.values()),
    }
    return daily,meta

def parse_1m_qa_from_source(source_path: Path, cal: dict[str,Any]) -> tuple[dict[date,DailyAcc],dict[str,Any]]:
    """Concatenate all 1min* sheets for measurement QA only."""
    member_name,raw,suffix=_workbook_payload(source_path)
    names=_sheet_names(raw,suffix)
    one=[n for n in names if str(n).strip().startswith("1min")]
    if not one:
        raise ValueError("no 1min sheet")
    merged: dict[date,DailyAcc]={}
    metas=[]
    for sh in one:
        d,m=parse_sheet_from_source(source_path,sh,cal)
        metas.append(m)
        for dt,acc in d.items():
            tgt=merged.setdefault(dt,DailyAcc([],[],[]))
            tgt.returns.extend(acc.returns)
            tgt.day_returns.extend(acc.day_returns)
            tgt.night_returns.extend(acc.night_returns)
            tgt.valid_bars+=acc.valid_bars
            tgt.invalid_bars+=acc.invalid_bars
            tgt.duplicate_keys+=acc.duplicate_keys
            for k,v in acc.session_counts.items():
                tgt.session_counts[k]=tgt.session_counts.get(k,0)+v
    return merged,{"workbook_member":member_name,"sheets":one,"components":metas}

def sumsq(xs:list[float])->float: return float(sum(x*x for x in xs))
def semipos(xs:list[float])->float: return float(sum(x*x for x in xs if x>=0))
def semineg(xs:list[float])->float: return float(sum(x*x for x in xs if x<0))

def pearson_corr(xs:list[float], ys:list[float]) -> float|None:
    if len(xs) != len(ys) or len(xs) < 2:
        return None
    mx=sum(xs)/len(xs); my=sum(ys)/len(ys)
    dx=[x-mx for x in xs]; dy=[y-my for y in ys]
    den=(sum(x*x for x in dx)*sum(y*y for y in dy))**0.5
    if den == 0:
        return None
    return float(sum(a*b for a,b in zip(dx,dy))/den)

def annual_files(folder:Path)->list[Path]:
    fs=list(folder.glob("N225microf_*.zip"))
    fs=sorted(fs,key=lambda p:p.name)
    if not fs: raise SystemExit("no 225Labo Micro annual minute files found")
    return fs

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--input-dir",type=Path,required=True)
    ap.add_argument("--calendar",type=Path,required=True)
    ap.add_argument("--output-dir",type=Path,required=True)
    ap.add_argument("--skip-1m-qa",action="store_true",help="Build the frozen 5m primary panel first; 1m measurement QA can be run separately.")
    ap.add_argument("--parser-commit",required=True,help="Git commit SHA of this exact parser version.")
    args=ap.parse_args()
    args.output_dir.mkdir(parents=True,exist_ok=True)
    cal=load_calendar(args.calendar)

    # Read annual 5m files from both legacy direct XLS (2006-2011) and later ZIP
    # containers, de-duplicating overlapping trading dates by choosing
    # the package whose nominal year equals the trading date year; if absent,
    # prefer the newest package deterministically.
    candidates: dict[date,list[tuple[int,Path,DailyAcc]]]={}
    sources=[]
    source_hash_lookup: dict[str,str]={}
    qa1m=[]
    for p in annual_files(args.input_dir):
        m=re.search(r"(20\d{2})",p.name)
        nominal=int(m.group(1)) if m else -1
        d5,meta5=parse_sheet_from_source(p,"5min",cal)
        if args.skip_1m_qa:
            qa1m.append({"file":p.name,"status":"QA_1M_DEFERRED_PRIMARY_5M_BUILD"})
        else:
            try:
                d1,meta1=parse_1m_qa_from_source(p,cal)
                # Aggregate-only measurement QA summary per file, no daily 1m export.
                common=sorted(set(d5)&set(d1))
                pairs=[]
                for d in common:
                    if d5[d].returns and d1[d].returns:
                        pairs.append((sumsq(d5[d].returns),sumsq(d1[d].returns)))
                if len(pairs)>=2:
                    xs=[a for a,b in pairs]; ys=[b for a,b in pairs]
                    corr=pearson_corr(xs,ys)
                    ratios=[b/a for a,b in pairs if a>0]
                    ratio_median=float(statistics.median(ratios)) if ratios else None
                else:
                    corr=None; ratio_median=None
                qa1m.append({"file":p.name,"common_days":len(pairs),"rv_1m_vs_5m_corr":corr,"rv_1m_to_5m_ratio_median":ratio_median,"meta":meta1})
            except Exception as exc:
                qa1m.append({"file":p.name,"status":"QA_1M_UNAVAILABLE","error":type(exc).__name__})
        for d,acc in d5.items():
            candidates.setdefault(d,[]).append((nominal,p,acc))
        source_digest=sha256_file(p)
        source_hash_lookup[p.name]=source_digest
        sources.append({
            "source_id":p.name,
            "sha256":source_digest,
            "size_bytes":p.stat().st_size,
            "nominal_year":nominal,
            "meta_5m":meta5,
        })

    selected={}
    overlap_days=0
    for d,items in candidates.items():
        if len(items)>1: overlap_days+=1
        exact=[x for x in items if x[0]==d.year]
        choice=(exact if exact else items)
        choice=sorted(choice,key=lambda x:(x[0],x[1].name),reverse=True)[0]
        selected[d]=choice

    rows=[]
    critical=[]
    for d,(nominal,p,acc) in sorted(selected.items()):
        if not acc.returns:
            critical.append({"trading_date":d.isoformat(),"issue":"NO_VALID_5M_RETURNS"})
            continue
        rv=sumsq(acc.returns); rp=semipos(acc.returns); rn=semineg(acc.returns)
        # Numerical identity is a hard check.
        if abs(rv-(rp+rn)) > max(1e-15,rv*1e-10):
            critical.append({"trading_date":d.isoformat(),"issue":"RV_RSV_IDENTITY_FAIL"})
        sched=schedule_for(d,cal)
        # Coverage ratio is relative to observed expected label grid implied by
        # the frozen session version, inclusive of endpoint/auction labels.
        expected=0
        for a,b in sched.get("day_session_segments",[])+sched.get("night_session_segments",[]):
            aa,bb=parse_hhmm(a),parse_hhmm(b)
            if "+1" in b: dur=(24*60-aa)+bb
            else: dur=bb-aa
            expected += dur//5 + 1
        coverage=min(1.0,acc.valid_bars/expected) if expected else 0.0
        rows.append({
            "trading_date":d.isoformat(),
            "rv_5m":rv,
            "rsv_pos_5m":rp,
            "rsv_neg_5m":rn,
            "day_session_rv":sumsq(acc.day_returns),
            "night_session_rv":sumsq(acc.night_returns),
            "n_5m_returns":len(acc.returns),
            "valid_5m_bars":acc.valid_bars,
            "session_coverage_ratio":coverage,
            "source_file_sha256":source_hash_lookup[p.name],
            "transform_version":"225LABO_MICRO_RVRSV_V1",
        })

    rows=sorted(rows,key=lambda r:r["trading_date"])
    panel=args.output_dir/"jnu_225labo_micro_daily_rvrsv_v1.csv"
    fieldnames=[
        "trading_date","rv_5m","rsv_pos_5m","rsv_neg_5m",
        "day_session_rv","night_session_rv","n_5m_returns",
        "valid_5m_bars","session_coverage_ratio","source_file_sha256",
        "transform_version"
    ]
    with panel.open("w",encoding="utf-8",newline="") as fh:
        writer=csv.DictWriter(fh,fieldnames=fieldnames,lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)
    panel_hash=sha256_file(panel)

    # Compact gap inventory.
    years_present=sorted({s["nominal_year"] for s in sources if s["nominal_year"]>0})
    missing_years=[y for y in range(2023,datetime.now().year+1) if y not in years_present]

    manifest={
        "version":"1.0",
        "source_license_classification":"225LABO_PERSONAL_USE_LOCAL_RAW_DERIVED_NON_RECONSTRUCTIVE_EXPORT",
        "raw_data_cloud_uploaded":False,
        "source_hashes":[{"source_id":s["source_id"],"sha256":s["sha256"]} for s in sources],
        "parser_version_commit":args.parser_commit,
        "calendar_session_version":cal.get("version"),
        "product_contract_coverage":{
            "venue":"OSE","product":"Nikkei 225 Micro Futures (JNU)","source_series":"225Labo annual JNU Micro central-contract minute packages",
            "years_present":years_present,"missing_year_packages":missing_years,
        },
        "date_range":[rows[0]["trading_date"] if rows else None,rows[-1]["trading_date"] if rows else None],
        "missingness_summary":{
            "selected_trading_days":len(rows),
            "annual_packages_missing":missing_years,
            "days_below_95pct_session_coverage":sum(1 for r in rows if r["session_coverage_ratio"]<0.95),
            "minimum_session_coverage_ratio":min((r["session_coverage_ratio"] for r in rows),default=None),
        },
        "duplicate_summary":{"overlap_trading_days_across_annual_packages":overlap_days},
        "derived_feature_definitions":{
            "rv_5m":"sum squared log close-to-close returns within each valid historical OSE session; no cross-session returns",
            "rsv_pos_5m":"sum squared non-negative 5m log returns",
            "rsv_neg_5m":"sum squared negative 5m log returns",
            "day_session_rv":"same measure restricted to DAY segments",
            "night_session_rv":"same measure restricted to NIGHT segments",
        },
        "derived_output_hash":panel_hash,
        "critical_data_quality_issues":critical,
        "measurement_qa_1m":qa1m,
        "measurement_qa_1m_deferred":bool(args.skip_1m_qa),
        "sources":sources,
    }
    mpath=args.output_dir/"jnu_225labo_micro_daily_rvrsv_v1_manifest.json"
    mpath.write_text(json.dumps(manifest,ensure_ascii=False,indent=2)+"\n",encoding="utf-8")
    print(json.dumps({
        "status":"DERIVED_PANEL_BUILT" if not critical else "DERIVED_PANEL_BUILT_WITH_CRITICAL_DQ",
        "panel":str(panel),"manifest":str(mpath),"rows":len(rows),
        "date_range":manifest["date_range"],"years_present":years_present,
        "missing_years":missing_years,"critical_issue_count":len(critical),
        "days_below_95pct_coverage":manifest["missingness_summary"]["days_below_95pct_session_coverage"],
        "derived_output_hash":panel_hash,
    },ensure_ascii=False,indent=2))

if __name__=="__main__":
    main()
