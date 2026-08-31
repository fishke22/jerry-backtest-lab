#!/usr/bin/env python3
"""Local-only OSE Mini/JNU intraday momentum-reversion sufficient-stat adapter.

Frozen family:
- H1 1-minute momentum: signal = sign(previous 1m return)
- H2 non-overlapping 10-minute mean reversion: signal = -sign(previous 10m return)

Reads personally licensed 225Labo raw minute workbooks locally and emits only
per-day non-reconstructive sufficient statistics. No interval return series or
raw prices are emitted.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import math
import re
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable

EXPECTED_HEADER=["日付","時間","始値","高値","安値","終値","出来高"]
TRANSFORM_VERSION="JNU_INTRADAY_MOMREV_G1_V1"

def sha256_file(path:Path)->str:
    h=hashlib.sha256()
    with path.open("rb") as fh:
        for b in iter(lambda:fh.read(1024*1024),b""):
            h.update(b)
    return h.hexdigest()

def normalize_date(v:Any)->date:
    if isinstance(v,datetime): return v.date()
    if isinstance(v,date): return v
    s=str(v).strip().replace("/","-")
    try: return datetime.fromisoformat(s).date()
    except ValueError: return datetime.strptime(s.split()[0],"%Y-%m-%d").date()

def normalize_time(v:Any)->time:
    if isinstance(v,datetime): return v.time().replace(microsecond=0)
    if isinstance(v,time): return v.replace(microsecond=0)
    if isinstance(v,(int,float)):
        sec=round(float(v)*86400)%86400
        return time(sec//3600,(sec%3600)//60,sec%60)
    return time.fromisoformat(str(v).strip()).replace(microsecond=0)

def minute_of_day(t:time)->int:
    return t.hour*60+t.minute

def parse_hhmm(s:str)->int:
    h,m=map(int,s.replace("+1","").split(":"))
    return h*60+m

def schedule_for(d:date,cal:dict[str,Any])->dict[str,Any]:
    ds=d.isoformat()
    for row in cal["ose_nikkei_index_futures"]:
        if ds<row["valid_from"]: continue
        if row["valid_to"] is not None and ds>row["valid_to"]: continue
        return row
    raise ValueError(f"no OSE calendar row for {ds}")

def iter_xlsx(raw:bytes,sheet:str)->Iterable[tuple[Any,...]]:
    import openpyxl
    wb=openpyxl.load_workbook(io.BytesIO(raw),read_only=True,data_only=True)
    try:
        ws=wb[sheet]
        yield from ws.iter_rows(values_only=True)
    finally:
        wb.close()

def iter_xls(raw:bytes,sheet:str)->Iterable[tuple[Any,...]]:
    import xlrd
    book=xlrd.open_workbook(file_contents=raw,on_demand=True)
    try:
        sh=book.sheet_by_name(sheet)
        for r in range(sh.nrows):
            vals=[]
            for c in range(sh.ncols):
                cell=sh.cell(r,c)
                if cell.ctype==xlrd.XL_CELL_DATE:
                    vals.append(xlrd.xldate.xldate_as_datetime(cell.value,book.datemode))
                else:
                    vals.append(cell.value)
            yield tuple(vals)
    finally:
        book.release_resources()

def workbook_payload(path:Path)->tuple[str,bytes,str]:
    if path.suffix.lower() in {".xls",".xlsx"}:
        return path.name,path.read_bytes(),path.suffix.lower()
    with zipfile.ZipFile(path) as zf:
        books=[i for i in zf.infolist() if not i.is_dir() and Path(i.filename).suffix.lower() in {".xls",".xlsx"}]
        if len(books)!=1:
            raise ValueError(f"{path.name}: expected exactly one workbook, got {len(books)}")
        info=books[0]
        return info.filename,zf.read(info),Path(info.filename).suffix.lower()

def sheet_names(raw:bytes,suffix:str)->list[str]:
    if suffix==".xlsx":
        import openpyxl
        wb=openpyxl.load_workbook(io.BytesIO(raw),read_only=True,data_only=True)
        try: return list(wb.sheetnames)
        finally: wb.close()
    import xlrd
    b=xlrd.open_workbook(file_contents=raw,on_demand=True)
    try: return list(b.sheet_names())
    finally: b.release_resources()

def find_header(rows:Iterable[tuple[Any,...]],source:str,sheet:str)->Iterable[tuple[Any,...]]:
    it=iter(rows)
    pre=[]
    for _ in range(12):
        try: row=next(it)
        except StopIteration: break
        vals=[str(x).strip() if x is not None else "" for x in list(row)[:7]]
        if vals==EXPECTED_HEADER:
            return it
        pre.append(vals)
    raise ValueError(f"{source}/{sheet}: standard header not found; preamble={pre[:3]}")

@dataclass
class DayBars:
    closes:dict[int,float]
    duplicate_rows:int
    invalid_rows:int
    def __init__(self):
        self.closes={}
        self.duplicate_rows=0
        self.invalid_rows=0

def source_files(folder:Path,product:str)->dict[int,Path]:
    if product=="MINI":
        files=list(folder.glob("N225minif_*.zip"))+list(folder.glob("225mini20*d.xls"))
    else:
        files=list(folder.glob("N225microf_*.zip"))
    out={}
    for p in files:
        m=re.search(r"(20\d{2})",p.name)
        if m: out[int(m.group(1))]=p
    return out

def parse_source_days(path:Path,start:date,end:date)->tuple[dict[date,DayBars],dict[str,Any]]:
    member,raw,suf=workbook_payload(path)
    names=sheet_names(raw,suf)
    one=[n for n in names if str(n).strip().startswith("1min")]
    if not one: raise ValueError(f"{path.name}: no 1min-prefixed sheet")
    days:dict[date,DayBars]={}
    relevant_rows=0
    for sh in one:
        rows=iter_xlsx(raw,sh) if suf==".xlsx" else iter_xls(raw,sh)
        it=find_header(rows,path.name,sh)
        for row in it:
            if not row or all(v in (None,"") for v in row): continue
            try: d=normalize_date(row[0])
            except Exception: continue
            if d<start or d>end: continue
            relevant_rows+=1
            day=days.setdefault(d,DayBars())
            try:
                t=normalize_time(row[1]); c=float(row[5])
            except Exception:
                day.invalid_rows+=1; continue
            if not (c>0 and math.isfinite(c)):
                day.invalid_rows+=1; continue
            m=minute_of_day(t)
            if m in day.closes:
                day.duplicate_rows+=1
                continue
            day.closes[m]=c
    meta={
        "workbook_member":member,
        "sheets":one,
        "relevant_rows":relevant_rows,
        "distinct_dates":len(days),
        "duplicate_rows":sum(x.duplicate_rows for x in days.values()),
        "invalid_rows":sum(x.invalid_rows for x in days.values())
    }
    return days,meta

def sign(x:float)->int:
    return 1 if x>0 else (-1 if x<0 else 0)

def expected_day_minutes(sched:dict[str,Any])->set[int]:
    out=set()
    for a,b in sched["day_session_segments"]:
        out.update(range(parse_hhmm(a),parse_hhmm(b)))
    return out

def day_stats(d:date,day:DayBars,cal:dict[str,Any])->dict[str,Any]:
    sched=schedule_for(d,cal)
    expected=expected_day_minutes(sched)
    observed=set(day.closes)&expected
    coverage=len(observed)/len(expected) if expected else 0.0

    h1_pairs=0; h1_acc_n=0; h1_correct=0; h1_payoff=0.0; h1_zero_prev=0; h1_zero_target=0
    h2_pairs=0; h2_acc_n=0; h2_correct=0; h2_payoff=0.0; h2_zero_prev=0; h2_zero_target=0
    h2_blocks_total=0

    for a,b in sched["day_session_segments"]:
        lo,hi=parse_hhmm(a),parse_hhmm(b)
        # Close-to-close 1m returns keyed by end-minute label.
        rets={}
        for t in range(lo+1,hi):
            if t-1 in day.closes and t in day.closes:
                rets[t]=math.log(day.closes[t]/day.closes[t-1])

        keys=sorted(rets)
        for i in range(1,len(keys)):
            if keys[i] != keys[i-1]+1:
                continue
            prev=rets[keys[i-1]]; target=rets[keys[i]]
            sp=sign(prev)
            if sp==0:
                h1_zero_prev+=1
                continue
            h1_pairs+=1
            h1_payoff+=sp*target
            st=sign(target)
            if st==0:
                h1_zero_target+=1
            else:
                h1_acc_n+=1
                if sp==st: h1_correct+=1

        # Non-overlapping 10m returns aligned to the first close-to-close return.
        blocks=[]
        t=lo+1
        while t+9 < hi:
            ks=list(range(t,t+10))
            if all(k in rets for k in ks):
                blocks.append(sum(rets[k] for k in ks))
            else:
                blocks.append(None)
            t+=10
        h2_blocks_total+=sum(x is not None for x in blocks)
        for i in range(1,len(blocks)):
            if blocks[i-1] is None or blocks[i] is None:
                continue
            prev=float(blocks[i-1]); target=float(blocks[i])
            sp=-sign(prev)
            if sp==0:
                h2_zero_prev+=1
                continue
            h2_pairs+=1
            h2_payoff+=sp*target
            st=sign(target)
            if st==0:
                h2_zero_target+=1
            else:
                h2_acc_n+=1
                if sp==st: h2_correct+=1

    return {
        "trading_date":d.isoformat(),
        "day_session_minute_coverage":coverage,
        "h1_pairs":h1_pairs,
        "h1_accuracy_denominator":h1_acc_n,
        "h1_correct":h1_correct,
        "h1_signal_payoff_sum":h1_payoff,
        "h1_daily_mean_signal_payoff":h1_payoff/h1_pairs if h1_pairs else None,
        "h1_zero_predecessor_count":h1_zero_prev,
        "h1_zero_target_count":h1_zero_target,
        "h2_nonoverlap_10m_blocks":h2_blocks_total,
        "h2_pairs":h2_pairs,
        "h2_accuracy_denominator":h2_acc_n,
        "h2_correct":h2_correct,
        "h2_signal_payoff_sum":h2_payoff,
        "h2_daily_mean_signal_payoff":h2_payoff/h2_pairs if h2_pairs else None,
        "h2_zero_predecessor_count":h2_zero_prev,
        "h2_zero_target_count":h2_zero_target,
    }

def main()->None:
    ap=argparse.ArgumentParser()
    ap.add_argument("--product",choices=["MINI","MICRO"],required=True)
    ap.add_argument("--input-dir",type=Path,required=True)
    ap.add_argument("--calendar",type=Path,required=True)
    ap.add_argument("--prereg",type=Path,required=True)
    ap.add_argument("--output-dir",type=Path,required=True)
    ap.add_argument("--parser-commit",required=True)
    args=ap.parse_args()
    args.output_dir.mkdir(parents=True,exist_ok=True)

    cal=json.loads(args.calendar.read_text(encoding="utf-8"))
    prereg=json.loads(args.prereg.read_text(encoding="utf-8"))
    stage="stage_a" if args.product=="MINI" else "stage_b"
    cfg=prereg["data"][stage]
    start=date.fromisoformat(cfg["date_from"])
    end=date.max if cfg["date_to"]=="latest available" else date.fromisoformat(cfg["date_to"])
    cov_min=float(prereg["data_quality"]["day_session_minute_label_coverage_minimum"])

    files=source_files(args.input_dir,args.product)
    candidates:dict[date,list[tuple[int,Path,DayBars]]]={}
    sources=[]
    critical=[]
    for year,p in sorted(files.items()):
        # Skip annual packages wholly outside fixed stage date range where possible.
        if year < start.year-1 or year > end.year+1:
            continue
        days,meta=parse_source_days(p,start,end)
        digest=sha256_file(p)
        sources.append({
            "source_id":p.name,
            "nominal_year":year,
            "sha256":digest,
            "size_bytes":p.stat().st_size,
            "meta_1m":meta
        })
        for d,bars in days.items():
            candidates.setdefault(d,[]).append((year,p,bars))

    rows=[]; excluded=[]; duplicate_overlap_days=0
    for d,items in sorted(candidates.items()):
        if len(items)>1: duplicate_overlap_days+=1
        exact=[x for x in items if x[0]==d.year]
        nominal,p,day=sorted(exact if exact else items,key=lambda x:(x[0],x[1].name),reverse=True)[0]
        if day.invalid_rows>0:
            critical.append({"trading_date":d.isoformat(),"issue":"INVALID_NONPOSITIVE_OR_UNPARSEABLE_CLOSE","count":day.invalid_rows})
            continue
        st=day_stats(d,day,cal)
        if st["day_session_minute_coverage"] < cov_min:
            excluded.append({
                "trading_date":d.isoformat(),
                "reason":"DAY_SESSION_MINUTE_COVERAGE_BELOW_GATE",
                "coverage":st["day_session_minute_coverage"]
            })
            continue
        if not st["h1_pairs"] or not st["h2_pairs"]:
            excluded.append({
                "trading_date":d.isoformat(),
                "reason":"NO_ELIGIBLE_SIGNAL_PAIRS",
                "h1_pairs":st["h1_pairs"],
                "h2_pairs":st["h2_pairs"]
            })
            continue
        st["source_file_sha256"]=sha256_file(p)
        st["transform_version"]=TRANSFORM_VERSION
        rows.append(st)

    stem="jnu_momrev_mini_stage_a_g1" if args.product=="MINI" else "jnu_momrev_micro_stage_b_g1"
    panel=args.output_dir/f"{stem}.csv"
    fields=[
        "trading_date","day_session_minute_coverage",
        "h1_pairs","h1_accuracy_denominator","h1_correct","h1_signal_payoff_sum","h1_daily_mean_signal_payoff","h1_zero_predecessor_count","h1_zero_target_count",
        "h2_nonoverlap_10m_blocks","h2_pairs","h2_accuracy_denominator","h2_correct","h2_signal_payoff_sum","h2_daily_mean_signal_payoff","h2_zero_predecessor_count","h2_zero_target_count",
        "source_file_sha256","transform_version"
    ]
    with panel.open("w",encoding="utf-8",newline="") as fh:
        w=csv.DictWriter(fh,fieldnames=fields,lineterminator="\n")
        w.writeheader(); w.writerows(rows)

    manifest={
        "version":"1.0",
        "candidate_id":prereg["candidate_id"],
        "stage":"A_TRUE_OSE_MINI" if args.product=="MINI" else "B_EXACT_JNU_MICRO",
        "source_license_classification":"225LABO_PERSONAL_USE_LOCAL_RAW_DERIVED_NON_RECONSTRUCTIVE_EXPORT",
        "raw_data_cloud_uploaded":False,
        "parser_version_commit":args.parser_commit,
        "calendar_session_version":cal.get("version"),
        "product_contract_coverage":{
            "venue":"OSE",
            "product":"Nikkei 225 Mini Futures" if args.product=="MINI" else "Nikkei 225 Micro Futures (JNU)",
            "date_from":cfg["date_from"],
            "date_to":cfg["date_to"]
        },
        "date_range":[rows[0]["trading_date"] if rows else None,rows[-1]["trading_date"] if rows else None],
        "missingness_summary":{
            "usable_trading_days":len(rows),
            "excluded_trading_days":len(excluded),
            "coverage_gate":cov_min
        },
        "duplicate_summary":{
            "annual_package_overlap_days":duplicate_overlap_days,
            "source_duplicate_minute_rows":sum(int(s["meta_1m"]["duplicate_rows"]) for s in sources)
        },
        "derived_feature_definitions":{
            "h1_daily_mean_signal_payoff":"daily mean of sign(previous 1m return) * next 1m return within continuous day-session segments",
            "h2_daily_mean_signal_payoff":"daily mean of -sign(previous non-overlap 10m return) * next non-overlap 10m return within continuous day-session segments",
            "h1_accuracy":"continuation accuracy among nonzero predecessor and nonzero target 1m pairs",
            "h2_accuracy":"reversal accuracy among nonzero predecessor and nonzero target adjacent 10m blocks"
        },
        "source_hashes":[{"source_id":s["source_id"],"sha256":s["sha256"]} for s in sources],
        "sources":sources,
        "excluded_days":excluded[:200],
        "critical_data_quality_issues":critical,
        "derived_output_hash":sha256_file(panel),
        "market_outcome_interpretation_performed":False
    }
    mp=args.output_dir/f"{stem}_manifest.json"
    mp.write_text(json.dumps(manifest,ensure_ascii=False,indent=2)+"\n",encoding="utf-8")
    print(json.dumps({
        "status":"DERIVED_PANEL_BUILT" if not critical else "DERIVED_PANEL_BUILT_WITH_CRITICAL_DQ",
        "product":args.product,
        "usable_days":len(rows),
        "excluded_days":len(excluded),
        "critical":len(critical),
        "date_range":manifest["date_range"],
        "panel_sha256":manifest["derived_output_hash"],
        "panel":str(panel),
        "manifest":str(mp)
    },ensure_ascii=False,indent=2))

if __name__=="__main__":
    main()
