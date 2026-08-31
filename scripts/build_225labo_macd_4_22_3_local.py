#!/usr/bin/env python3
from __future__ import annotations

import argparse,csv,hashlib,io,json,math,re,zipfile
from datetime import date,datetime,time
from pathlib import Path
from typing import Any,Iterable

EXPECTED_HEADER=["日付","時間","始値","高値","安値","終値","出来高"]
TRANSFORM_VERSION="JNU_MACD_4_22_3_POSTPUBLICATION_G1_V1"

def sha256_file(path:Path)->str:
    h=hashlib.sha256()
    with path.open('rb') as fh:
        for b in iter(lambda:fh.read(1024*1024),b''):h.update(b)
    return h.hexdigest()

def norm_date(v:Any)->date:
    if isinstance(v,datetime):return v.date()
    if isinstance(v,date):return v
    s=str(v).strip().replace('/','-')
    try:return datetime.fromisoformat(s).date()
    except:return datetime.strptime(s.split()[0],'%Y-%m-%d').date()

def norm_time(v:Any)->time:
    if isinstance(v,datetime):return v.time().replace(microsecond=0)
    if isinstance(v,time):return v.replace(microsecond=0)
    if isinstance(v,(int,float)):
        sec=round(float(v)*86400)%86400;return time(sec//3600,(sec%3600)//60,sec%60)
    return time.fromisoformat(str(v).strip()).replace(microsecond=0)

def minute(t:time)->int:return t.hour*60+t.minute

def hhmm(s:str)->int:
    h,m=map(int,s.replace('+1','').split(':'));return h*60+m

def schedule_for(d:date,cal:dict[str,Any])->dict[str,Any]:
    ds=d.isoformat()
    for r in cal['ose_nikkei_index_futures']:
        if ds<r['valid_from']:continue
        if r['valid_to'] is not None and ds>r['valid_to']:continue
        return r
    raise ValueError(ds)

def workbook_payload(path:Path):
    if path.suffix.lower() in {'.xls','.xlsx'}:return path.name,path.read_bytes(),path.suffix.lower()
    with zipfile.ZipFile(path) as zf:
        books=[i for i in zf.infolist() if not i.is_dir() and Path(i.filename).suffix.lower() in {'.xls','.xlsx'}]
        if len(books)!=1:raise ValueError(f'{path.name}: workbook count {len(books)}')
        i=books[0];return i.filename,zf.read(i),Path(i.filename).suffix.lower()

def sheet_names(raw:bytes,suf:str):
    if suf=='.xlsx':
        import openpyxl
        wb=openpyxl.load_workbook(io.BytesIO(raw),read_only=True,data_only=True)
        try:return list(wb.sheetnames)
        finally:wb.close()
    import xlrd
    b=xlrd.open_workbook(file_contents=raw,on_demand=True)
    try:return list(b.sheet_names())
    finally:b.release_resources()

def iter_xlsx(raw:bytes,sh:str)->Iterable[tuple[Any,...]]:
    import openpyxl
    wb=openpyxl.load_workbook(io.BytesIO(raw),read_only=True,data_only=True)
    try:yield from wb[sh].iter_rows(values_only=True)
    finally:wb.close()

def iter_xls(raw:bytes,sh:str)->Iterable[tuple[Any,...]]:
    import xlrd
    b=xlrd.open_workbook(file_contents=raw,on_demand=True)
    try:
        s=b.sheet_by_name(sh)
        for r in range(s.nrows):
            vals=[]
            for c in range(s.ncols):
                cell=s.cell(r,c);vals.append(xlrd.xldate.xldate_as_datetime(cell.value,b.datemode) if cell.ctype==xlrd.XL_CELL_DATE else cell.value)
            yield tuple(vals)
    finally:b.release_resources()

def find_header(rows,source,sh):
    it=iter(rows)
    for _ in range(12):
        row=next(it,None)
        if row is None:break
        vals=[str(x).strip() if x is not None else '' for x in list(row)[:7]]
        if vals==EXPECTED_HEADER:return it
    raise ValueError(f'{source}/{sh}: header')

def source_files(folder:Path,product:str)->dict[int,Path]:
    fs=(list(folder.glob('N225minif_*.zip'))+list(folder.glob('225mini20*d.xls'))) if product=='MINI' else list(folder.glob('N225microf_*.zip'))
    out={}
    for p in fs:
        m=re.search(r'(20\d{2})',p.name)
        if m:out[int(m.group(1))]=p
    return out

def parse_daily_close(path:Path,start:date,end:date,cal:dict[str,Any]):
    member,raw,suf=workbook_payload(path);one=[n for n in sheet_names(raw,suf) if str(n).strip().startswith('1min')]
    if not one:raise ValueError(f'{path.name}: no 1min')
    days={};dup=invalid=0
    for sh in one:
        rows=iter_xlsx(raw,sh) if suf=='.xlsx' else iter_xls(raw,sh);it=find_header(rows,path.name,sh)
        for row in it:
            if not row or all(v in (None,'') for v in row):continue
            try:d=norm_date(row[0])
            except:continue
            if d<start or d>end:continue
            try:m=minute(norm_time(row[1]));c=float(row[5])
            except:invalid+=1;continue
            if not(c>0 and math.isfinite(c)):invalid+=1;continue
            sched=schedule_for(d,cal);active=[]
            for a,b in sched['day_session_segments']:active.extend(range(hhmm(a),hhmm(b)))
            if m not in set(active):continue
            rec=days.setdefault(d,{})
            if m in rec:dup+=1;continue
            rec[m]=c
    out={}
    for d,bars in days.items():
        sched=schedule_for(d,cal);active=[]
        for a,b in sched['day_session_segments']:active.extend(range(hhmm(a),hhmm(b)))
        expected=set(active);coverage=len(expected&set(bars))/len(expected) if expected else 0
        last=max(active) if active else None
        if last in bars:out[d]={'close':bars[last],'coverage':coverage}
    return out,{'member':member,'sheets':one,'duplicate_rows':dup,'invalid_rows':invalid,'distinct_dates':len(out)}

def ema(values:list[float],n:int)->list[float]:
    alpha=2/(n+1);out=[];e=None
    for x in values:
        e=x if e is None else alpha*x+(1-alpha)*e;out.append(e)
    return out

def moving_block_prob(xs:list[float],block:int,samples:int,seed:int):
    import random
    rng=random.Random(seed);n=len(xs);starts=list(range(max(1,n-block+1)));means=[]
    for _ in range(samples):
        z=[]
        while len(z)<n:
            s=rng.choice(starts);z.extend(xs[s:s+block])
        means.append(sum(z[:n])/n)
    ys=sorted(means)
    def q(p):
        pos=(len(ys)-1)*p;lo=int(pos);hi=min(lo+1,len(ys)-1);f=pos-lo;return ys[lo]*(1-f)+ys[hi]*f
    return {'prob_mean_positive':sum(x>0 for x in means)/samples,'ci95':[q(.025),q(.975)]}

def main():
    ap=argparse.ArgumentParser();ap.add_argument('--product',choices=['MINI','MICRO'],required=True);ap.add_argument('--input-dir',type=Path,required=True);ap.add_argument('--calendar',type=Path,required=True);ap.add_argument('--prereg',type=Path,required=True);ap.add_argument('--output-dir',type=Path,required=True);ap.add_argument('--parser-commit',required=True);args=ap.parse_args();args.output_dir.mkdir(parents=True,exist_ok=True)
    cal=json.loads(args.calendar.read_text(encoding='utf-8'));pre=json.loads(args.prereg.read_text(encoding='utf-8'));st='stage_a' if args.product=='MINI' else 'stage_b';cfg=pre[st]
    eval_start=date.fromisoformat(cfg['evaluation_from']);eval_end=date.max if cfg['evaluation_to']=='latest available' else date.fromisoformat(cfg['evaluation_to'])
    warm_start=date(eval_start.year-1,1,1);files=source_files(args.input_dir,args.product);all_days={};sources=[];source_hash={};critical=[]
    for year,p in sorted(files.items()):
        if year<warm_start.year or year>eval_end.year:continue
        dd,meta=parse_daily_close(p,warm_start,eval_end,cal);digest=sha256_file(p);source_hash[p.name]=digest;sources.append({'source_id':p.name,'nominal_year':year,'sha256':digest,'meta_1m':meta})
        if meta['invalid_rows']>0:critical.append({'year':year,'issue':'INVALID_ROWS','count':meta['invalid_rows']})
        for d,v in dd.items():
            if d not in all_days or year==d.year:all_days[d]=(p,v)
    dates=sorted(all_days);prices=[all_days[d][1]['close'] for d in dates];e4=ema(prices,4);e22=ema(prices,22);macd=[a-b for a,b in zip(e4,e22)];sig=ema(macd,3);hist=[a-b for a,b in zip(macd,sig)]
    raw_signal=[1 if h>0 else (-1 if h<0 else 0) for h in hist];implemented=[];pos=0
    for i in range(len(dates)):
        if i>=1 and raw_signal[i-1]!=0:pos=raw_signal[i-1]
        implemented.append(pos)
    rows=[];excluded=[]
    for i,d in enumerate(dates):
        if d<eval_start or d>eval_end or i==0:continue
        cov=all_days[d][1]['coverage']
        if cov<float(pre['data_quality']['day_session_minute_label_coverage_minimum']):excluded.append({'date':d.isoformat(),'reason':'COVERAGE','coverage':cov});continue
        r=math.log(prices[i]/prices[i-1]);p=implemented[i];pay=p*r if p else 0.0;den=1 if p and r!=0 else 0;correct=1 if den and ((p>0 and r>0) or (p<0 and r<0)) else 0
        rows.append({'trading_date':d.isoformat(),'raw_signal_at_prior_close':raw_signal[i-1],'implemented_position':p,'underlying_close_to_close_return':r,'strategy_return':pay,'accuracy_denominator':den,'correct':correct,'day_session_minute_coverage':cov,'source_file_sha256':source_hash[all_days[d][0].name],'transform_version':TRANSFORM_VERSION})
    stem='jnu_macd_4_22_3_mini_stage_a_g1' if args.product=='MINI' else 'jnu_macd_4_22_3_micro_stage_b_g1';panel=args.output_dir/f'{stem}.csv';fields=list(rows[0].keys()) if rows else []
    with panel.open('w',encoding='utf-8',newline='') as fh:
        w=csv.DictWriter(fh,fieldnames=fields,lineterminator='\n');w.writeheader();w.writerows(rows)
    manifest={'version':'1.0','candidate_id':pre['candidate_id'],'stage':'A_TRUE_OSE_MINI' if args.product=='MINI' else 'B_EXACT_JNU_MICRO','source_license_classification':'225LABO_PERSONAL_USE_LOCAL_RAW_DERIVED_NON_RECONSTRUCTIVE_EXPORT','raw_data_cloud_uploaded':False,'parser_version_commit':args.parser_commit,'calendar_session_version':cal.get('version'),'product_contract_coverage':{'venue':'OSE','product':'Nikkei 225 Mini Futures' if args.product=='MINI' else 'Nikkei 225 Micro Futures (JNU)','evaluation_from':cfg['evaluation_from'],'evaluation_to':cfg['evaluation_to']},'date_range':[rows[0]['trading_date'] if rows else None,rows[-1]['trading_date'] if rows else None],'missingness_summary':{'usable_trading_days':len(rows),'excluded_trading_days':len(excluded)},'duplicate_summary':{'source_duplicate_minute_rows':sum(x['meta_1m']['duplicate_rows'] for x in sources)},'derived_feature_definitions':{'raw_signal_at_prior_close':'sign(MACD(4,22,3) histogram at prior day close)','implemented_position':'prior-close signal implemented one close later','strategy_return':'implemented_position * daily close-to-close log return'},'source_hashes':[{'source_id':x['source_id'],'sha256':x['sha256']} for x in sources],'sources':sources,'excluded_days':excluded[:100],'critical_data_quality_issues':critical,'derived_output_hash':sha256_file(panel),'market_outcome_interpretation_performed':False}
    mp=args.output_dir/f'{stem}_manifest.json';mp.write_text(json.dumps(manifest,ensure_ascii=False,indent=2)+'\n',encoding='utf-8');print(json.dumps({'usable_days':len(rows),'excluded_days':len(excluded),'critical':len(critical),'date_range':manifest['date_range'],'panel_sha256':manifest['derived_output_hash'],'panel':str(panel),'manifest':str(mp)},ensure_ascii=False,indent=2))
if __name__=='__main__':main()
